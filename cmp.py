import pandas as pd
import numpy as np

# ── 1. Load both sheets ───────────────────────────────────────────────────────
FILE = "/Users/shishir/Downloads/Daily TPM P&L Report Keystone vs Kortex validation (1).xlsx"

print(f"\n[1/6] Loading data from '{FILE}'...")
ks_raw = pd.read_excel(FILE, sheet_name="keystone output 2023 onwards")
print(f"      ✓ Keystone loaded  : {len(ks_raw):,} rows x {len(ks_raw.columns)} cols")

kx_raw = pd.read_excel(FILE, sheet_name="v3_output_kortex(6th april)")
print(f"      ✓ Kortex loaded    : {len(kx_raw):,} rows x {len(kx_raw.columns)} cols")

# ── 2. Normalise column names (lowercase, strip spaces) ──────────────────────
print("\n[2/6] Normalising column names and join keys...")
ks_raw.columns = ks_raw.columns.str.strip().str.lower()
kx_raw.columns = kx_raw.columns.str.strip().str.lower()

# ── 3. Define join keys (normalised) ─────────────────────────────────────────
JOIN_KEYS = ["plan_to_nbr", "plan_to_nm", "fisc_yr", "fisc_pd", "bu", "catg"]
# Kortex uses different names for some keys — rename to match Keystone
kx_raw = kx_raw.rename(columns={
    "bu_cd":        "bu",
    "altn_catg_nm": "catg",
})

# Strip leading zeros from plan_to_nbr so 0000300003 == 300003, etc.
for df in [ks_raw, kx_raw]:
    df["plan_to_nbr"] = df["plan_to_nbr"].astype(str).str.lstrip("0")
    df["plan_to_nm"]  = df["plan_to_nm"].astype(str).str.strip().str.upper()
    df["bu"]          = df["bu"].astype(str).str.strip().str.upper()
    df["catg"]        = df["catg"].astype(str).str.strip().str.upper()
    df["fisc_yr"]     = df["fisc_yr"].astype(str).str.strip()
    df["fisc_pd"]     = df["fisc_pd"].astype(str).str.strip().str.zfill(2)

print(f"      ✓ Join keys        : {JOIN_KEYS}")
print(f"      ✓ Keystone unique grain rows : {ks_raw[JOIN_KEYS].drop_duplicates().shape[0]:,}")
print(f"      ✓ Kortex unique grain rows   : {kx_raw[JOIN_KEYS].drop_duplicates().shape[0]:,}")

# ── 4. Column mapping: Keystone name → Kortex name ───────────────────────────
# Only columns that exist in both sources (after normalisation)
print("\n[3/6] Mapping metric columns between Keystone and Kortex...")
COL_MAP = {
    "crm_le_orig_gros_sales_mntly": "sum(crm_le_orig_gross_sales)",
    "ibp_gsv":                      "sum(ibp_gsv)",
    "lp_ibp_gsv":                   "sum(lp_ibp_gsv)",
    "ibp_kmf":                      "sum(ibp_kmf)",
    "crm_base_kmf_oi_bb":           "sum(crm_base_kmf_oi_bb)",
    "bud_gsv":                      "sum(bud_gsv)",
    "bud_kmf":                      "sum(bud_kmf)",
    "se_gsv":                       "sum(se_gsv)",
    "se_kmf":                       "sum(se_kmf)",
    "se3_kmf":                      "sum(se3_kmf)",
    "crm_le_cop":                   "sum(crm_le_cop)",
    "yago_crm_le_cop":              "sum(yr_ago_crm_le_cop)",
    "crm_le_lbs":                   "sum(crm_le_lbs)",
    "yago_crm_le_lbs":              "sum(yr_ago_crm_le_lbs)",
    "cop_se":                       "sum(cop_se)",
    "cop_bud":                      "sum(cop_budj)",
    "se3_gsv":                      "sum(se3_gsv)",
    "lp_ibp_kmf":                   "sum(lp_ibp_kmf)",   # not in KS sample but map anyway
    "bud_kmf_adj":                  "sum(bud_kmf_adj)",
    "se_kmf_adj":                   "sum(se_kmf_adj)",
    "yago_gsv":                     "sum(yr_ago_gsv)",
    "yago_kmf":                     "sum(yr_ago_kmf)",
}

# Keep only columns that actually exist in each dataframe
ks_metric_cols = [k for k in COL_MAP if k in ks_raw.columns]
kx_metric_cols = [COL_MAP[k] for k in ks_metric_cols if COL_MAP[k] in kx_raw.columns]
ks_metric_cols = [k for k in ks_metric_cols if COL_MAP[k] in kx_raw.columns]

# Subset and rename Kortex metric cols to match Keystone names
ks = ks_raw[JOIN_KEYS + ks_metric_cols].copy()
kx = kx_raw[JOIN_KEYS + kx_metric_cols].copy()
kx = kx.rename(columns={COL_MAP[k]: k for k in ks_metric_cols})

# ── 5. Full outer join ────────────────────────────────────────────────────────
print("\n[4/6] Performing full outer join on grain keys...")
merged = pd.merge(
    ks, kx,
    on=JOIN_KEYS,
    how="outer",
    suffixes=("_ks", "_kx"),
    indicator=True
)

# Rename _merge indicator to something readable
merged["source"] = merged["_merge"].map({
    "left_only":  "Keystone only",
    "right_only": "Kortex only",
    "both":       "Both",
})
merged.drop(columns="_merge", inplace=True)
ks_only  = (merged["source"] == "Keystone only").sum()
kx_only  = (merged["source"] == "Kortex only").sum()
matched  = (merged["source"] == "Both").sum()
print(f"      ✓ Total rows after join : {len(merged):,}")
print(f"        ├─ Matched (both)     : {matched:,}")
print(f"        ├─ Keystone only      : {ks_only:,}")
print(f"        └─ Kortex only        : {kx_only:,}")

# ── 6. Compute absolute diff and % diff for every metric ─────────────────────
print(f"\n[5/6] Computing abs diff and % diff for {len(ks_metric_cols)} metrics...")
for col in ks_metric_cols:
    ks_col  = f"{col}_ks"
    kx_col  = f"{col}_kx"
    merged[ks_col] = pd.to_numeric(merged[ks_col], errors="coerce")
    merged[kx_col] = pd.to_numeric(merged[kx_col], errors="coerce")

    merged[f"{col}_abs_diff"] = merged[kx_col] - merged[ks_col]
    merged[f"{col}_pct_diff"] = np.where(
        merged[ks_col].notna() & (merged[ks_col] != 0),
        (merged[kx_col] - merged[ks_col]) / merged[ks_col] * 100,
        np.nan
    )

# ── 7. Summary: rows with any difference ─────────────────────────────────────
abs_diff_cols = [c for c in merged.columns if c.endswith("_abs_diff")]
merged["has_diff"] = (
    merged[abs_diff_cols].abs().fillna(0) > 0.01
).any(axis=1)

only_diffs = merged[merged["has_diff"]].copy()
print(f"      ✓ Rows with at least one metric difference : {len(only_diffs):,}")
print(f"      ✓ Rows with no differences                 : {(~merged['has_diff']).sum():,}")

# ── 8. Write to Excel output ──────────────────────────────────────────────────
OUT = "/Users/shishir/Downloads/comparison_output.xlsx"
print(f"\n[6/6] Writing results to '{OUT}'...")

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    # Sheet 1: Full comparison (all rows)
    merged.drop(columns="has_diff").to_excel(
        writer, sheet_name="Full_Comparison", index=False
    )
    print(f"      ✓ Sheet 'Full_Comparison'   written ({len(merged):,} rows)")
    
    # Sheet 2: Only rows with differences
    only_diffs.drop(columns="has_diff").to_excel(
        writer, sheet_name="Differences_Only", index=False
    )
    print(f"      ✓ Sheet 'Differences_Only'  written ({len(only_diffs):,} rows)")

    # Sheet 3: Summary stats per metric
    summary = pd.DataFrame({
        "metric":          ks_metric_cols,
        "ks_sum":          [merged[f"{c}_ks"].sum() for c in ks_metric_cols],
        "kx_sum":          [merged[f"{c}_kx"].sum() for c in ks_metric_cols],
        "total_abs_diff":  [merged[f"{c}_abs_diff"].abs().sum() for c in ks_metric_cols],
        "rows_with_diff":  [(merged[f"{c}_abs_diff"].abs().fillna(0) > 0.01).sum()
                            for c in ks_metric_cols],
        "avg_pct_diff":    [merged[f"{c}_pct_diff"].mean() for c in ks_metric_cols],
    })
    summary.to_excel(writer, sheet_name="Metric_Summary", index=False)
    print(f"      ✓ Sheet 'Metric_Summary'    written ({len(summary)} metrics)")

    # ── Apply basic formatting ──────────────────────────────────────────────
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    DIFF_FILL = PatternFill("solid", start_color="FFF2CC")

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        # Bold headers
        for cell in ws[1]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)
        # Highlight diff cols yellow
        if sheet_name in ("Full_Comparison", "Differences_Only"):
            for cell in ws[1]:
                if cell.value and ("_diff" in str(cell.value) or "source" in str(cell.value)):
                    cell.fill = PatternFill("solid", start_color="FFD700")
                    cell.font = Font(bold=True, name="Arial", size=10)

print(f"✅  Done → {OUT}")
print(f"   Total rows after join : {len(merged)}")
print(f"   Keystone-only rows    : {(merged['source']=='Keystone only').sum()}")
print(f"   Kortex-only rows      : {(merged['source']=='Kortex only').sum()}")
print(f"   Matched rows          : {(merged['source']=='Both').sum()}")
print(f"   Rows with differences : {merged['has_diff'].sum()}")